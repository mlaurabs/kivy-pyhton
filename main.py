from os.path import exists

import openpyxl
from kivymd.app import MDApp
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.core.window import Window
import os
from kivymd.uix.pickers import MDDatePicker
from kivymd.uix.dialog import MDDialog
from kivy.utils import get_color_from_hex
from kivymd.uix.button import MDRaisedButton, MDRectangleFlatButton
import locale
# this sets the date time formats to pt_PT, there are many other options for currency, numbers etc.
# You might need to install language package  "sudo apt-get install language-pack-pt-base" pt means Portuguese
locale.setlocale(locale.LC_TIME, 'pt_BR.utf8') # dessa forma, o date picker é traduzido para português Brasil

Window.size = (350, 600)


# acessando id de um componente da tela:  self.ids.(nome do id).text(Se o text não for colocado, será apenas o endereço e não o valor do componente)

class SplashScreen(Screen):
    pass

class Home(Screen):

    def on_pre_enter(self, *args):# ao gerar a tela, a função é executada

        #verifica se já há um arquivo de dados salvo
        if not exists('Data\dados_base.xlsx'): # se não houver um arquivo criado
            # criar o arquivo
            print("Vou criar o arquivo")
            dataBase = openpyxl.Workbook()
            dataBase_sheet = dataBase['Sheet']
            dataBase_sheet.append(['Valor', 'Data', 'Para', 'Desc'])
            dataBase.create_sheet("Saldo")
            dataBase_saldo = dataBase["Saldo"]
            dataBase_saldo.append(['Saldo'])
            dataBase_saldo.append([float(0.00)])

            # deixando em negrito os headers da tabela
            bold = openpyxl.styles.Font(bold=True)  # guardamos o estilo em uma variável
            for row in dataBase_sheet["A1:E1"]: # adicionamos esse estilo para cada célula
                for cell in row:
                    cell.font = bold
            dataBase_saldo['A1'].font = bold

            dataBase.save('Data\dados_base.xlsx')

            """
            linha = dataBase_saldo.max_row # pega a última linha /célula com valor --> saldo atual

            if(dataBase_saldo[f"A{linha}"].value < 0): # se o saldo estiver negativo
                print("entrei aqui")
                self.ids.saldo.text_color = get_color_from_hex("#FF0000")

            self.ids.saldo.text = 'R$ ' + f"{dataBase_saldo[f'A{linha}'].value: .2f}"  # dessa forma o saldo é apresentado com 2 casas decimais
            dataBase.save('Data\dados_base.xlsx')
            dataBase.close()
            """
        else: # se já houver um arquivo criado
            # carregar e abrir o arquivo
            print("O arquivo já foi criado")
            dataBase = openpyxl.load_workbook('Data\dados_base.xlsx')
            dataBase_saldo = dataBase["Saldo"]

        # Atualiza o valor do saldo na home screen
        linha = dataBase_saldo.max_row  # pega a última linha /célula com valor --> saldo atual
        print(linha)

        if (dataBase_saldo[f"A{linha}"].value < 0): # se o saldo estiver negativo
            print("entrei aqui")
            self.ids.saldo.text_color = get_color_from_hex("#FF0000")

        self.ids.saldo.text = 'R$ ' + f"{dataBase_saldo[f'A{linha}'].value: .2f}" # dessa forma o saldo é apresentado com 2 casas decimais
        dataBase.close() # fechar o arquivo

    def chama(self):
        print("entrei no chama")
        os.system("Data\dados_base.xlsx")


"""
book_base = openpyxl.Workbook()  # creating the excel file we'll be using
book_base_sheet = book_base['Sheet'] #selecting our sheet
book_base_sheet.append(['Data', 'Tanque', 'Leitura']) # adding to our sheet data (each word before the comma is a column)
book_base.save('dados_base.xlsx') # saving and giving the file a name
book_base.close() # closing our workbook


ft = Font(bold=True)
>>> for row in ws["A1:C1"]:
...     for cell in row:
...         cell.font = ft


#editando a planilha base
            edit_base = openpyxl.load_workbook('dados_base.xlsx')
            edit_base_sheet = edit_base['Sheet']
            edit_base_sheet.append([self.date_read, self.termo_tanque, self.valor_read])
            edit_base.save('dados_base.xlsx')
            edit_base.close()

"""


class Transferir(Screen):

    dialog = True

    def TrocaFormato(self):
        # Em português usamos vírgula para separar casas decimais, mas em inglês o ponto é utilizado
        # Logo, para que o valor escrito com vígula possa ser formatado para float corretamente, é necessário trocar a vírgula pelo ponto
        valor = self.ids.valor.text
        virgula = ","
        ponto = "."

        if virgula in valor:  # se tiver vírgula no valor...
            return float(valor.replace(virgula, ponto))  # trocar por ponto
        else:
            return float(valor)

    def Home(self):
        print("entrei na Home")
        dataBase = openpyxl.load_workbook('Data\dados_base.xlsx')
        dataBase_saldo = dataBase["Saldo"]


        linha = dataBase_saldo.max_row  # pega a última linha /célula com valor

        if (dataBase_saldo[f"A{linha}"].value < 0):
            print("entrei aqui")
            self.manager.get_screen("home").ids.saldo.text_color = get_color_from_hex("#FF0000")    # manager.get_screen(nome da screen) é utlizado para que possamos acessar um componente de outra tela
        else:
            self.manager.get_screen("home").ids.saldo.text_color = get_color_from_hex("#F9F9F9")

        self.manager.get_screen("home").ids.saldo.text = 'R$ ' + f"{dataBase_saldo[f'A{linha}'].value: .2f}"  # dessa forma o saldo é apresentado com 2 casas decimais
        self.manager.current = 'home'

        dataBase.close()

    def SaveInfo(self, obj):

        self.dialog.dismiss()  # it closes the alert showed

        dataBase = openpyxl.load_workbook('Data\dados_base.xlsx')
        dataBase_saldo = dataBase["Saldo"]
        dataBase_sheet = dataBase["Sheet"]

        valor = self.TrocaFormato()

        para = str(self.ids.para.text)
        data = str(self.ids.data.text)
        desc = str(self.ids.desc.text)

        linha = dataBase_saldo.max_row
        saldo = float(dataBase_saldo[f'A{linha}'].value)  # saldo mais recente

        saldo_atual = (saldo - valor)
        print(saldo_atual)
        dataBase_saldo.append([saldo_atual])

        dataBase_sheet.append([valor, data, para, desc])
        dataBase.save('Data\dados_base.xlsx')
        dataBase.close()

        self.Home()

    def continuar(self, obj):  # in case the user dismisses de alert given when he says NO
        self.dialog.dismiss()  # it closes the alert showed
        self.manager.current = 'home'  # directs to the first registering

    def close_dialog(self, obj):  # closes the alert
        self.dialog.dismiss()

    def mostrar_dialogo(self):  # code nedded for the dialog to be showed

        valor = self.TrocaFormato()
        print(valor)

        dataBase = openpyxl.load_workbook('Data\dados_base.xlsx')
        dataBase_saldo = dataBase["Saldo"]

        linha = dataBase_saldo.max_row
        saldo = float(dataBase_saldo[f'A{linha}'].value)  # saldo mais recente
        dataBase.close()

        if(valor > saldo):
            if self.dialog:
                self.dialog = MDDialog(
                    title="Saldo Insuficiente!",
                    text="Se decidir continuar com esta ação, sua conta estará negativada. O valor será descontado eventualmente de sua conta.",
                    buttons=[
                        MDRectangleFlatButton(text="Continuar", on_press=self.continuar, on_release=self.SaveInfo),
                        MDRaisedButton(text="Cancelar", on_press=self.close_dialog)
                    ]
                )
                self.dialog.open()
        else:
            self.SaveInfo(self.dialog)



    def ShowingDate(self, instance, value, date_range):
        self.date = str(value) # get the date
        # formatando a data
        self.data_br = self.date.split('-')
        self.ids.data.text = self.data_br[2] + '/' + self.data_br[1] + '/' + self.data_br[0]  # shows it in text

    def Cancelar(self, instance, date_range):
        self.ids.data.text = "Selecione uma data"

    def show_date_picker(self):
        # estilando o date picker
        date_dialog = MDDatePicker(title="Selecione uma data", primary_color= get_color_from_hex("#002171"),
        accent_color= get_color_from_hex("#FFFFFF"),
        selector_color= get_color_from_hex("#5472D3"),
        text_toolbar_color= get_color_from_hex("#FFFFFF"),
        text_weekday_color= get_color_from_hex("#002171"),
        text_current_color= get_color_from_hex("#FFFFF"),
        text_button_color= get_color_from_hex("#002171"))
        date_dialog.bind(on_save=self.ShowingDate, on_cancel=self.Cancelar)
        date_dialog.open()

class Depositar(Screen):

    def Home(self):
        print("entrei na Home")

        dataBase = openpyxl.load_workbook('Data\dados_base.xlsx')
        dataBase_saldo = dataBase["Saldo"]

        linha = dataBase_saldo.max_row  # pega a última linha /célula com valor

        if (dataBase_saldo[f"A{linha}"].value < 0):
            print("entrei aqui")
            self.manager.get_screen("home").ids.saldo.text_color = get_color_from_hex("#FF0000")    # manager.get_screen(nome da screen) é utlizado para que possamos acessar um componente de outra tela
        else:
            self.manager.get_screen("home").ids.saldo.text_color = get_color_from_hex("#F9F9F9")

        self.manager.get_screen("home").ids.saldo.text = 'R$ ' + f"{dataBase_saldo[f'A{linha}'].value: .2f}"  # dessa forma o saldo é apresentado com 2 casas decimais
        self.manager.current = 'home'

        dataBase.close()


    def SaveInfo(self):
        dataBase = openpyxl.load_workbook('Data\dados_base.xlsx')
        dataBase_sheet = dataBase['Sheet']
        dataBase_saldo = dataBase['Saldo']

        # Em português usamos vírgula para separar casas decimais, mas em inglês o ponto é utilizado
        # Logo para que o valor escrito com vígula possa ser formatado para float corretamente, é necessário trocar a vírgula pelo ponto
        valor = self.ids.valor.text
        virgula = ","
        ponto = "."

        if virgula in valor: # se tiver vírgula no valor...
            valor = valor.replace(virgula, ponto)  # trocar por ponto

        data = str(self.ids.data.text)

        linha = dataBase_saldo.max_row
        saldo = float(dataBase_saldo[f'A{linha}'].value)  # saldo mais recente

        self.saldo_atual = (saldo + float(valor))
        print(self.saldo_atual)
        dataBase_saldo.append([self.saldo_atual])

        dataBase_sheet.append([valor, data, None, None])
        dataBase.save('Data\dados_base.xlsx')
        dataBase.close()

        self.Home()


    def ShowingDate(self, instance, value, date_range):
        self.date = str(value) # get the date
        # formatando a data
        self.data_br = self.date.split('-')
        self.ids.data.text = self.data_br[2] + '/' + self.data_br[1] + '/' + self.data_br[0]  # shows it in text

    def Cancelar(self, instance, date_range):
        self.ids.data.text = "Selecione uma data"

    def show_date_picker(self):
        # estilando o date picker
        date_dialog = MDDatePicker(title="Selecione uma data", primary_color= get_color_from_hex("#002171"),
        accent_color= get_color_from_hex("#FFFFFF"),
        selector_color= get_color_from_hex("#5472D3"),
        text_toolbar_color= get_color_from_hex("#FFFFFF"),
        text_weekday_color= get_color_from_hex("#002171"),
        text_current_color= get_color_from_hex("#FFFFF"),
        text_button_color= get_color_from_hex("#002171"))
        date_dialog.bind(on_save=self.ShowingDate, on_cancel=self.Cancelar)
        date_dialog.open()

class MiniBanco(MDApp):

    def build(self):
        sm = ScreenManager()
        sm.add_widget(SplashScreen(name="splashScreen"))
        sm.add_widget(Home(name="home"))
        sm.add_widget(Transferir(name="transferir"))
        sm.add_widget(Depositar(name="depositar"))

        return sm

if __name__ == "__main__":
    MiniBanco().run()

"""
Tasks for back-end:

create a function to generate the excel file DONE
create a function to verify if there is already an excel file on the directory DONE
update the balance DONE
create a function to save data into the excel file DONE
create a function to go through all the file´s data and show it on the recycler view(component)
create a function to open the excel file DONE
create the Transferir functionality DONE
create the Depositar functionality DONE

caso a conta fique negativada DONE
o saldo fica em vermelho
um alert aparece quando tentar transferir além do valor na conta
quando for depositado um valor, será descontado dele
 
"""



